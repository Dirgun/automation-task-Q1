# -*- coding: utf-8 -*-
"""
Created on Wed Jan 17 17:02:24 2024

@author: turzo
"""

# automation.py
# Import necessary libraries
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime
import time
from datetime import datetime, timedelta
import openpyxl
from openpyxl import workbook

# open workbook
path_to_excel_file = 'C:\\Users\ASUS\OneDrive\Desktop\Q1\Q1.xlsx'
wb = openpyxl.load_workbook(path_to_excel_file)
# select by sheet name
now = datetime.now()
ws = wb[now.strftime("%A")]

# Initialize the Chrome driver
path_to_chromedriver = 'C:\\Users\ASUS\OneDrive\Desktop\Q1\chromedriver.exe'
service = Service(path_to_chromedriver)
driver = webdriver.Chrome(service=service)

def automation():
    for i in range(2,9):
        # Navigate to Google
        url = "https://www.google.com"
        driver.get(url)
        print(str(i))

        # Find the search box, enter the keyword and submit
        search_box = driver.find_element(By.NAME,'q')
        search_box.send_keys(ws['B'+str(i)].value)

        time.sleep(2)
    
        suggestions = driver.find_elements(By.CLASS_NAME, "OBMEnb")

        search_results_text = [result.text for result in suggestions]

        # for replacing \n's
        split_list = [s.split('\n') for s in search_results_text]
        # pop empty string
        split_list.pop()

        # putting the item with 0th index in a variable
        my_list = split_list[0]

        # converting list into a string
        my_string = ', '.join(my_list)

        # splitting the list into individual items
        my_list2 = my_string.split(',')
 
        # shortest string
        shortest_string = min(my_list2, key=len)
        print("Shortest String",shortest_string)

        # longest string
        longest_string = max(my_list2, key=len)
        print("Longest String",longest_string)

        ws['D'+str(i)] = longest_string
        ws['E'+str(i)] = shortest_string
        wb.save(path_to_excel_file)
        
# Get the current time
now = datetime.now()
# Get current hour and minute
now = datetime.now()
current_hour = now.hour
current_minute = now.minute

while True:
    now = datetime.now()
    if now.hour == current_hour and now.minute == current_minute:
        automation()
        time.sleep(60)






# -*- coding: utf-8 -*-
"""
Created on Wed Jan 17 17:02:24 2024

@author: turzo
"""
# automation.py
# Import necessary libraries
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import datetime
import openpyxl
from openpyxl import workbook

# open workbook
wb = openpyxl.load_workbook('C:\\Users\ASUS\OneDrive\Desktop\Q1.xlsx')
# delete current sheet
if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']
 
data = [('keyword1', 'Dhaka'),
        ('keyword2', 'University'),
        ('keyword3', 'Cricket'),
        ('keyword4', 'Bombay'),
        ('keyword5', 'Football'),
        ('keyword6', 'Paper'),
        ('keyword7', 'Knife'),]   
 
days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

# create new sheets named maonday, tuesday, ...
for day in days:
    ws = wb.create_sheet(day)
# add the data
    for row in data:
        ws.append(row)
        
# select by sheet name
ws = wb['Monday']

# Initialize the Chrome driver
service = Service('C:\\Users\ASUS\.spyder-py3\chromedriver.exe')
driver = webdriver.Chrome(service=service)

def automation():
    for i in range(1,8):
        # Navigate to Google
        url = "https://www.google.com"
        driver.get(url)
        print(str(i))

        # Find the search box, enter the keyword and submit
        search_box = driver.find_element(By.NAME,'q')
        search_box.send_keys(ws['B'+str(i)].value)

        time.sleep(2)
    
        suggestions = driver.find_elements(By.CLASS_NAME, "OBMEnb")

        search_results_text = [result.text for result in suggestions]

        # for replacing \n's
        split_list = [s.split('\n') for s in search_results_text]
        # pop empty string
        split_list.pop()

        # putting the item with 0th index in a variable
        my_list = split_list[0]

        # converting list into a string
        my_string = ', '.join(my_list)

        # splitting the list into individual items
        my_list2 = my_string.split(',')
 
        # shortest string
        shortest_string = min(my_list2, key=len)
        print("Shortest String",shortest_string)

        # longest string
        longest_string = max(my_list2, key=len)
        print("Longest String",longest_string)

        ws['D'+str(i)] = longest_string
        ws['E'+str(i)] = shortest_string
        wb.save('C:\\Users\ASUS\OneDrive\Desktop\Q1.xlsx')
        
run_at = datetime.time(hour=17, minute=56)

while True:
    # Get the current time
    now = datetime.datetime.now().time()

    # Check if the current time is equal to the specified time
    if now.hour == run_at.hour and now.minute == run_at.minute:
        automation()
        
        time.sleep(60)  # Delay execution for 60 seconds